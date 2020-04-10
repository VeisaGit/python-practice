from telegram import Update, ParseMode, Bot
from telegram.ext import Updater, CallbackContext
from telegram.ext import Filters, MessageHandler
from Commands import *
from db import *
from openpyxl import Workbook

def main():
    print('start')
    updater = Updater(
        token='1093115241:AAHC4pU5yX1w9jhPRkcXEqv9Wy_REo8C8Mc',
        use_context=True,
    )

    updater.dispatcher.add_handler(MessageHandler(filters=Filters.all, callback=message_handler))

    updater.start_polling() #делает работу бота постоянной
    updater.idle()


def message_handler(update: Update, context: CallbackContext):
    bot = Bot
    text = update.effective_message.text
    user = update.effective_user

    if text == '/instructions':
        update.message.reply_text(instructions(text), parse_mode=ParseMode.HTML)

    elif text == '/start' or text == '/bot':
        update.message.reply_text(information_about_bot(text), parse_mode=ParseMode.HTML)

    elif text == '/average':
        #TODO: вынести этот в commands.py, в отдельную ф-ю
        monthly_average = monthly_average_query(user_id=user.id)
        sistolic_average = monthly_average[0][0]
        diastolic_average = monthly_average[0][1]
        update.message.reply_text('Среднее значение вашего давления \n за последние 30 дней -  <b>{}/{}</b>'.format(round(sistolic_average), round(diastolic_average)), parse_mode=ParseMode.HTML)
        # TODO: Если среднее значение отклоняется от нормы, сообщить об этом.

    elif text == '/export30':
        # TODO: вынести этот в commands.py, в отдельную ф-ю
        monthly_statistic = monthly_statistic_query(user_id=user.id)
        text = '{}'.format(monthly_statistic)
        # update.message.reply_text(text=text)TODO: нужно было для отражение работы ф-ии.

        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        ws['A1'] = 'Систолическое'
        ws['B1'] = 'Диасталическое'


        for row in monthly_statistic:
            ws.append(row)

        # Save the file TODO: в отчет добавить дату.
        wb.save("report_30_days.xlsx")

        # update.message.reply_text(chat_id) TODO: нужно было для отражение работы ф-ии.
        doc = open('report_30_days.xlsx', 'rb')
        update.message.reply_document(doc)

    else:
        # TODO: вынести этот в commands.py, в отдельную ф-ю
        pressure_value = processing_user_input(text)
        systolic_pressure_value = int(pressure_value[0])
        diastolic_pressure_value = int(pressure_value[1])

        update.message.reply_text(successful_user_message(user), parse_mode=ParseMode.HTML)

        add_message_to_db(
            user_id=user.id,
            systolic_pressure=systolic_pressure_value,
            diastolic_pressure=diastolic_pressure_value,
            )
        # TODO: Рассмотреть возможность внесения комментария к факту фиксации давления (например указать самочувствие - отпразить это в описании к предложению внести коммент)
        # TODO: Отдельно включить краткую инструкцию в виде описания/картинки как правильно мерить давление.
        # TODO: Рассмотреть возможность добавление команды - самое высокое значение, самое низкое значение.


if __name__ == '__main__':
    main()

