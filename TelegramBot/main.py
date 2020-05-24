# - *- coding: utf- 8 - *-
from telegram import ParseMode
from telegram.ext import Updater, CallbackContext, ConversationHandler
from telegram.ext import Filters, MessageHandler, CommandHandler
from openpyxl import Workbook
from Commands import *
from const import TG_TOKEN, TG_URL
import logging, uuid, os, time

ex_mark = u'\U00002757'  # exclamation mark smile
up_red_triangle = u'\U0001F53A'  # up-pointing red triangle
down_red_triangle = u'\U0001F53B'  # down-pointing red triangle

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger('info_logger ')


def main():
    print(logger.info('bot start '))
    bot = Updater(TG_TOKEN, TG_URL, use_context=True)

    bot.dispatcher.add_handler(
        ConversationHandler(entry_points=[MessageHandler(Filters.regex('/average'), start_average_query)],
                            # запускает данный диалог, точка входа.
                            # шаги диалога, у каждого шага есть название и обработчик на который этот шаг реагирует.
                            states={
                                "time_limit": [MessageHandler(Filters.regex('Да'), query_with_heart_rate),
                                               MessageHandler(Filters.regex('Нет'), query_without_heart_rate)],
                                "amount_of_days_pressure_without_heart_rate": [
                                    MessageHandler(Filters.regex('24 часа'), pressure_query_without_heart_rate_1),
                                    MessageHandler(Filters.regex('7 дней'), pressure_query_without_heart_rate_7),
                                    MessageHandler(Filters.regex('14 дней'), pressure_query_without_heart_rate_14),
                                    MessageHandler(Filters.regex('30 дней'), pressure_query_without_heart_rate_30)],
                                "amount_of_days_pressure_with_heart_rate": [
                                    MessageHandler(Filters.regex('24 часа'), pressure_query_with_heart_rate_1),
                                    MessageHandler(Filters.regex('7 дней'), pressure_query_with_heart_rate_7),
                                    MessageHandler(Filters.regex('14 дней'), pressure_query_with_heart_rate_14),
                                    MessageHandler(Filters.regex('30 дней'), pressure_query_with_heart_rate_30)],
                                # TODO: реализовать: если пользователь ничего не вводит какое-то время, то заканчивать сессию.
                            },
                            # выход из диалога или можно использовать при некорректном вводе информации пользователем.
                            fallbacks=[MessageHandler(
                                Filters.text | Filters.video | Filters.photo | Filters.document, wrong_input)]
                            )
    )

    bot.dispatcher.add_handler(MessageHandler(filters=Filters.all, callback=message_handler))

    bot.start_polling()  # делает работу бота постоянной
    bot.idle()


def message_handler(update: Update, context: CallbackContext):
    text = update.effective_message.text
    user = update.effective_user

    if text == '/instructions':
        update.message.reply_text(instructions(text), parse_mode=ParseMode.HTML)

    elif text == '/start' or text == '/bot':
        update.message.reply_text(information_about_bot(text), parse_mode=ParseMode.HTML)

    elif text == '/id':
        request_id_quantity = len(unique_id_counter())
        request_id_quantity2 = len(unique_id_counter2())
        update.message.reply_text(f"{request_id_quantity}_{request_id_quantity2}", parse_mode=ParseMode.HTML)

    elif text == '/del':
        if user_id_exists_in_db(user_id=user.id):
            del_last_insert(user_id=user.id)
            update.message.reply_text("Последняя запись удалена.")
        else:
            update.message.reply_text("Ваших записей нет в базе данных.")

    elif text == '/export30':
        # TODO: вынести этот в commands.py, в отдельную ф-ю
        monthly_statistic = monthly_statistic_query(user_id=user.id)

        # TODO: возможно проблема недоступности при одновременном обращении к файлу excel решается через указание, если эта книга не доступна, использовать другую.
        # При сохранении файла вставить в разрез имени генерацию случайного числа, чтобы файл генерился уникальный. ИЛИ лушче дату.

        # Export to excel file
        wb = Workbook()
        ws = wb.active

        ws['A1'] = 'Дата и время(МСК) измерения'
        ws['B1'] = 'Систолическое'
        ws['C1'] = 'Диасталическое'
        ws['D1'] = 'Пульс'

        for row in monthly_statistic:
            ws.append(row)

        ws.column_dimensions["A"].width = 25

        random_id = uuid.uuid1()

        excel_file_name = 'report_30_days_' + str(random_id.hex) + '.xlsx'

        wb.save(excel_file_name)

        doc = open(excel_file_name, 'rb')
        update.message.reply_document(doc)

        time.sleep(3)

        file_path = '' + excel_file_name
        os.remove(file_path)


    else:
        # TODO: вынести этот в commands.py, в отдельную ф-ю
        pressure_value = processing_user_input(text)
        deviation_from_normal_pressure_and_heart_rate = thirty_day_average_query_with_heart_rate(user_id=user.id)
        percent_of_deviation = 0.2
        percent = ' на 20% '

        if len(pressure_value) == 2:
            systolic_pressure_value = int(pressure_value[0])
            diastolic_pressure_value = int(pressure_value[1])
            heart_rate = 0

        elif len(pressure_value) == 3:
            systolic_pressure_value = int(pressure_value[0])
            diastolic_pressure_value = int(pressure_value[1])
            heart_rate = int(pressure_value[2])

        def deviation_pressure():
            systolic_deviation = int(deviation_from_normal_pressure_and_heart_rate[0][0])
            diastolic_deviation = int(deviation_from_normal_pressure_and_heart_rate[0][1])

            if systolic_pressure_value > systolic_deviation + (systolic_deviation * percent_of_deviation) \
                    or diastolic_pressure_value > diastolic_deviation + (diastolic_deviation * percent_of_deviation):
                update.message.reply_text(
                    f'{ex_mark}Введенные вами данные <u>давления</u> выше{up_red_triangle}{percent}, '
                    f'чем средний показатель за последние 30 дней. Пожалуйста, '
                    f'обратите на это внимание.', parse_mode=ParseMode.HTML)

            elif systolic_pressure_value < systolic_deviation - (systolic_deviation * percent_of_deviation) \
                    or diastolic_pressure_value < diastolic_deviation - (diastolic_deviation * percent_of_deviation):
                update.message.reply_text(
                    f'{ex_mark}Введенные вами данные <u>давления</u> ниже{down_red_triangle}{percent}, '
                    f'чем средний показатель за последние 30 дней. Пожалуйста, '
                    f'обратите на это внимание.', parse_mode=ParseMode.HTML)

        def deviation_heart_rate():
            heart_deviation = int(deviation_from_normal_pressure_and_heart_rate[0][2])

            if heart_rate > heart_deviation + (heart_deviation * percent_of_deviation):
                update.message.reply_text(
                    f'{ex_mark}Введенные вами данные <u>пульса</u> выше{up_red_triangle}{percent},'
                    f' чем средний показатель за последние 30 дней.'
                    ' Пожалуйста, обратите на это внимание.', parse_mode=ParseMode.HTML)

            elif 0 < heart_rate < heart_deviation - (heart_deviation * percent_of_deviation):
                update.message.reply_text(
                    f'{ex_mark}Введенные вами данные <u>пульса</u> ниже{down_red_triangle}{percent}, '
                    f'чем средний показатель за последние 30 дней.'
                    ' Пожалуйста, обратите на это внимание.', parse_mode=ParseMode.HTML)


        add_message_to_db(
            user_id=user.id,
            systolic_pressure=systolic_pressure_value,
            diastolic_pressure=diastolic_pressure_value,
            heart_rate=heart_rate,
        )


        # Проверка данных пользователя на отклонение от среднего значения
        try:
            deviation_pressure()
            deviation_heart_rate()
        except:
            pass

        # Вывод сообщения об успешности внесения данных в БД
        update.message.reply_text(successful_user_message(user), parse_mode=ParseMode.HTML)

        # update.message.reply_text(
        #     "Предложения и замечания по работе бота можете написать в нашем <a href = \'https://t.me/sdpl_bot_chat\'>чате </a>",
        #     parse_mode=ParseMode.HTML)



        # TODO: Рассмотреть возможность внесения комментария к факту фиксации давления (например указать самочувствие - отпразить это в описании к предложению внести коммент).
        #  Он будет вводить через наажтие кнопки , которая появляется вместе с ответным сообщением о том, что запись успешно добавлена в дневник.
        # TODO: Отдельно включить краткую инструкцию в виде описания/картинки как правильно мерить давление.
        # TODO: Рассмотреть возможность добавление команды - самое высокое значение, самое низкое значение.


if __name__ == '__main__':
    main()
